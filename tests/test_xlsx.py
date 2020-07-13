import io

import pytest
from openpyxl import Workbook, load_workbook

from docci.xlsx import xlsx_to_bytes, xlsx_to_file, dicts_to_xlsx, xlsx_from_bytes, xlsx_from_file


@pytest.fixture()
def xlsx() -> Workbook:
    xlsx_ = Workbook()
    xlsx_.active.append(["sam"])
    return xlsx_


def test_xlsx_to_bytes_saves_xlsx_as_valid_bytes(xlsx: Workbook) -> None:
    bytes_ = xlsx_to_bytes(xlsx)
    stream = io.BytesIO(bytes_)
    actual_xlsx = load_workbook(stream)

    assert tuple(actual_xlsx.active.values) == tuple(xlsx.active.values)


def test_xlsx_to_file_saves_xlsx_as_file_attachment(xlsx: Workbook) -> None:
    xlsx_file = xlsx_to_file(xlsx, "sam.xlsx")

    assert xlsx_file.name == "sam.xlsx"
    assert xlsx_file.content == xlsx_to_bytes(xlsx)


def test_dicts_to_xlsx() -> None:
    dicts = [{"col1": "val1", "col2": "val2"}, {"col1": "val3", "col2": "val4"}]
    xlsx = dicts_to_xlsx(dicts)
    assert tuple(xlsx.active.values) == (("col1", "col2"), ("val1", "val2"), ("val3", "val4"))


def test_dicts_to_xlsx_with_custom_headers() -> None:
    dicts = [{"col1": "val1", "col2": "val2"}, {"col1": "val3", "col2": "val4"}]
    xlsx = dicts_to_xlsx(dicts, headers=["Column 1", "Column 2"])
    assert tuple(xlsx.active.values) == (
        ("Column 1", "Column 2"),
        ("val1", "val2"),
        ("val3", "val4"),
    )


def test_xlsx_from_bytes(xlsx: Workbook) -> None:
    actual_xlsx = xlsx_from_bytes(xlsx_to_bytes(xlsx))

    assert tuple(actual_xlsx.active.values) == tuple(xlsx.active.values)


def test_xlsx_from_file(xlsx: Workbook) -> None:
    actual_xlsx = xlsx_from_file(xlsx_to_file(xlsx, "sam.xlsx"))

    assert tuple(actual_xlsx.active.values) == tuple(xlsx.active.values)
